# -*- coding: utf-8 -*-
# QR/Barcode Attendance (Tkinter) — write to Google Sheets, keep original logic
# Requirements: opencv-contrib-python, gspread, google-auth, openpyxl, pillow

import os, re, io
from datetime import datetime, time
import threading

# ---- GUI & Local roster loader (unchanged behavior) ----
import tkinter as tk
from tkinter import messagebox
from tkinter import font as tkfont
from PIL import Image, ImageTk  # สำหรับไฟล์ .jpg/.png
from openpyxl import load_workbook  # ใช้เฉพาะอ่าน roster เดิมจาก students.xlsx

# ---- Camera / Barcode decode (OpenCV contrib) ----
import cv2
import numpy as np

# ---- Google Sheets (write logs instead of Excel) ----
import gspread
from google.oauth2.service_account import Credentials

# ===================== utils: วัน/ช่วง และชื่อชีต roster =====================
WEEKDAY_EN_TO_TH = {
    "Monday": "จันทร์",
    "Tuesday": "อังคาร",
    "Wednesday": "พุธ",
    "Thursday": "พฤหัสบดี",
    "Friday": "ศุกร์",
    "Saturday": "เสาร์",
    "Sunday": "อาทิตย์",
}

def get_session_th_and_cutoff(now_t: time):
    """คืนค่า (session_th, cutoff_time) -> ('เช้า'|'บ่าย', cutoff)"""
    if now_t < time(12, 0, 0):
        return "เช้า", time(9, 10, 0)
    return "บ่าย", time(13, 10, 0)

def roster_sheet_name_for_now(now_dt: datetime) -> str:
    """ชื่อชีตใน students.xlsx เช่น 'จันทร์เช้า', 'อังคารบ่าย'"""
    day_th = WEEKDAY_EN_TO_TH[now_dt.strftime("%A")]
    session_th, _ = get_session_th_and_cutoff(now_dt.time())
    return f"{day_th}{session_th}"

def load_roster_id_and_seat(xlsx_path: str, sheet_name: str):
    """
    อ่าน students.xlsx ชีต {sheet_name}
    คาดหวังคอลัมน์: A=Student ID, B=Full Name, C=Seat  (แถว 1 เป็นหัวคอลัมน์)
    คืน:
      roster_by_id   : { sid -> (full_name, seat) }
      roster_by_seat : { seat -> (sid, full_name) }
    """
    try:
        wb = load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        tk.Tk().withdraw()
        messagebox.showerror("students.xlsx not found", str(e))
        raise SystemExit

    if sheet_name not in wb.sheetnames:
        tk.Tk().withdraw()
        messagebox.showerror("Roster sheet missing", f"'students.xlsx' has no sheet named '{sheet_name}'")
        raise SystemExit

    ws = wb[sheet_name]
    roster_by_id, roster_by_seat = {}, {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        sid, full_name, seat = (row[:3] if row else (None, None, None))
        if not sid:
            continue
        sid = str(sid).strip()
        full_name = "" if full_name is None else str(full_name).strip()
        seat = "" if seat is None else str(seat).strip()
        roster_by_id[sid] = (full_name, seat)
        if seat:
            roster_by_seat[seat.upper()] = (sid, full_name)
    return roster_by_id, roster_by_seat

def normalize_seat(s: str) -> str:
    return (s or "").strip().upper()

def extract_student_id(raw_input):
    """Extract digits then interpret by length."""
    digits = re.findall(r'\d+', raw_input or "")
    if digits:
        d = ''.join(digits)
        if len(d) == 14:
            return d[3:13]    # Some QR/barcode formats
        elif len(d) == 10:
            return d          # Normal student ID
        elif 1 <= len(d) <= 3:
            return d          # Test short id
    return None

# ===================== เวลาและวันปัจจุบัน =====================
now = datetime.now()
today_str = now.strftime("%Y-%m-%d")
current_time = now.time()
day_of_week_en = now.strftime("%A")     # Monday, Tuesday, ...
day_of_week = " " + day_of_week_en

# ===================== หยุดทำงานถ้าเป็นวันเสาร์หรืออาทิตย์ =====================
if day_of_week_en in ["Saturday", "Sunday"]:
    tk.Tk().withdraw()
    messagebox.showinfo("Weekend", f"Today is {day_of_week_en}. Attendance is closed.")
    raise SystemExit

# ===================== กำหนดช่วงเวลา/เส้นตัดสาย =====================
session_th, cutoff_time = get_session_th_and_cutoff(current_time)  # 'เช้า' หรือ 'บ่าย'
session_en = "Morning" if session_th == "เช้า" else "Afternoon"   # สำหรับสถานะ

# ===================== โหลดชื่อจาก students.xlsx (ชีตตามวัน+ช่วงภาษาไทย) =====================
roster_sheet = roster_sheet_name_for_now(now)
try:
    ROSTER_BY_ID, ROSTER_BY_SEAT = load_roster_id_and_seat("students.xlsx", roster_sheet)
except SystemExit:
    raise
except Exception as e:
    tk.Tk().withdraw()
    messagebox.showerror("Roster load failed", str(e))
    raise SystemExit

# ===================== Google Sheets: เตรียมเชื่อมต่อ & ชีตของวัน =====================
# ตั้งค่าจาก ENV หรือแก้ตรงนี้ให้เป็นของคุณ
SHEET_ID = os.getenv("GOOGLE_SHEET_ID", "").strip()    # <-- ใส่ Spreadsheet ID
SERVICE_ACCOUNT_JSON = os.getenv("GOOGLE_APPLICATION_CREDENTIALS", "").strip()  # path ไป JSON

if not SHEET_ID:
    tk.Tk().withdraw()
    messagebox.showerror("Google Sheets", "Please set GOOGLE_SHEET_ID environment variable.")
    raise SystemExit

creds = None
try:
    if SERVICE_ACCOUNT_JSON and os.path.exists(SERVICE_ACCOUNT_JSON):
        # โหลดจากไฟล์ path
        creds = Credentials.from_service_account_file(
            SERVICE_ACCOUNT_JSON,
            scopes=[
                "https://www.googleapis.com/auth/spreadsheets",
                "https://www.googleapis.com/auth/drive.readonly",
            ],
        )
    else:
        tk.Tk().withdraw()
        messagebox.showerror("Google Sheets",
                             "Please set GOOGLE_APPLICATION_CREDENTIALS to your service account JSON path.")
        raise SystemExit
    gclient = gspread.authorize(creds)
    gsheet = gclient.open_by_key(SHEET_ID)
except Exception as e:
    tk.Tk().withdraw()
    messagebox.showerror("Google Sheets auth/open failed", str(e))
    raise SystemExit

# worksheet ชื่อเป็นวันที่ (เหมือนเดิมกับ Excel ที่ใช้ sheet_name = today_str)
GS_SHEET_TITLE = today_str

def open_or_create_worksheet(sh: gspread.Spreadsheet, title: str):
    try:
        ws = sh.worksheet(title)
    except gspread.exceptions.WorksheetNotFound:
        ws = sh.add_worksheet(title=title, rows=2000, cols=10)
        # สร้างหัวตารางให้เหมือนเดิม
        ws.append_row([f"Attendance date {today_str} | Session {session_en}"], value_input_option="USER_ENTERED")
        ws.append_row(["Student ID", "Full Name", "Seat", "Check-in Time", "Status"], value_input_option="USER_ENTERED")
    return ws

try:
    ws_gs = open_or_create_worksheet(gsheet, GS_SHEET_TITLE)
except Exception as e:
    tk.Tk().withdraw()
    messagebox.showerror("Google Sheets worksheet", str(e))
    raise SystemExit

# set กันซ้ำ: ทั้งจากรหัส และจากที่นั่ง (โหลดจาก Google Sheets)
scanned_ids = set()
used_seats = set()
try:
    values = ws_gs.get_all_values()
    # แถว 1-2 เป็น header; เริ่มข้อมูลจริงที่แถว 3
    for r in values[2:]:
        if not r:
            continue
        sid = (str(r[0]).strip() if len(r) > 0 and r[0] else "")
        seat = (str(r[2]).strip() if len(r) > 2 and r[2] else "")
        if sid:
            scanned_ids.add(sid)
        if seat:
            used_seats.add(normalize_seat(seat))
except Exception:
    # ถ้าอ่านไม่ได้ ไม่ต้องล้มโปรแกรม
    pass

# ===================== ฟังก์ชันหลัก: เช็กชื่อ (บันทึกลง Google Sheets) =====================
def _save_row(sid: str, full_name: str, seat: str):
    now = datetime.now()
    time_str = now.strftime("%H:%M:%S")
    status = "On time" if now.time() <= cutoff_time else "Late"

    # Append ลง Google Sheet (แทน Excel)
    try:
        ws_gs.append_row([sid, full_name, seat, time_str, status], value_input_option="USER_ENTERED")
    except Exception as e:
        messagebox.showerror("Google Sheets append failed", str(e))
        return

    scanned_ids.add(sid)
    if seat:
        used_seats.add(normalize_seat(seat))

    result_label.config(
        text=f"✅ {sid} ({full_name})  Seat: {seat or '-'}  |  {time_str}  ({status})",
        fg="navy"
    )

def check_in(event=None):
    global session_th, session_en, cutoff_time, ROSTER_BY_ID, ROSTER_BY_SEAT, roster_sheet, scanned_ids, used_seats, ws_gs

    raw_id = entry.get().strip()
    raw_seat = seat_entry.get().strip()
    entry.delete(0, tk.END)
    seat_entry.delete(0, tk.END)

    # อัปเดต session/roster ถ้าข้ามเที่ยงหรือข้ามวัน (รวมอัปเดตชีต Google Sheet ด้วย)
    now = datetime.now()
    new_session_th, new_cutoff = get_session_th_and_cutoff(now.time())
    new_session_en = "Morning" if new_session_th == "เช้า" else "Afternoon"
    new_sheet = roster_sheet_name_for_now(now)

    if new_sheet != roster_sheet:
        session_th, session_en, cutoff_time = new_session_th, new_session_en, new_cutoff
        roster_sheet = new_sheet
        try:
            ROSTER_BY_ID, ROSTER_BY_SEAT = load_roster_id_and_seat("students.xlsx", roster_sheet)
        except SystemExit:
            return
        except Exception as e:
            messagebox.showerror("Roster load failed", str(e))
            return
        # เปลี่ยน worksheet ตามวันที่ใหม่
        new_title = now.strftime("%Y-%m-%d")
        try:
            ws_gs = open_or_create_worksheet(gsheet, new_title)
        except Exception as e:
            messagebox.showerror("Google Sheets worksheet switch failed", str(e))
            return
        # รีโหลดกันซ้ำจากชีตใหม่
        try:
            scanned_ids.clear(); used_seats.clear()
            values = ws_gs.get_all_values()
            for r in values[2:]:
                if not r:
                    continue
                sid = (str(r[0]).strip() if len(r) > 0 and r[0] else "")
                seat = (str(r[2]).strip() if len(r) > 2 and r[2] else "")
                if sid: scanned_ids.add(sid)
                if seat: used_seats.add(normalize_seat(seat))
        except Exception:
            pass

    sid = extract_student_id(raw_id) if raw_id else None
    seat_input = normalize_seat(raw_seat) if raw_seat else None

    # ต้องมีอย่างน้อยอย่างใดอย่างหนึ่ง
    if not sid and not seat_input:
        messagebox.showwarning("Missing input", "Please enter either a Student ID or a Seat (or both).")
        return

    # Case 1: ID only
    if sid and not seat_input:
        if sid not in ROSTER_BY_ID:
            messagebox.showwarning("Not in roster", f"Student ID {sid} is not in sheet '{roster_sheet}'.")
            return
        full_name, seat_from_table = ROSTER_BY_ID[sid]
        seat_from_table = normalize_seat(seat_from_table)

        if sid in scanned_ids:
            messagebox.showwarning("Duplicate", f"{sid} has already checked in.")
            return
        if seat_from_table and seat_from_table in used_seats:
            messagebox.showwarning("Seat taken", f"Seat {seat_from_table} is already used.")
            return

        _save_row(sid, full_name, seat_from_table)
        return

    # Case 2: Seat only
    if seat_input and not sid:
        if seat_input not in ROSTER_BY_SEAT:
            messagebox.showwarning("Seat not found", f"Seat {seat_input} is not in sheet '{roster_sheet}'.")
            return
        sid_from_seat, full_name = ROSTER_BY_SEAT[seat_input]

        if sid_from_seat in scanned_ids:
            messagebox.showwarning("Duplicate", f"{sid_from_seat} has already checked in.")
            return
        if seat_input in used_seats:
            messagebox.showwarning("Seat taken", f"Seat {seat_input} is already used.")
            return

        _save_row(sid_from_seat, full_name, seat_input)
        return

    # Case 3: Both provided -> must match roster
    if sid and seat_input:
        if sid not in ROSTER_BY_ID:
            messagebox.showwarning("Not in roster", f"Student ID {sid} is not in sheet '{roster_sheet}'.")
            return
        full_name, seat_from_table = ROSTER_BY_ID[sid]
        seat_from_table = normalize_seat(seat_from_table)

        if seat_from_table and seat_input != seat_from_table:
            messagebox.showwarning("Seat mismatch", f"{sid} is assigned to seat {seat_from_table} (not {seat_input}).")
            return

        if sid in scanned_ids:
            messagebox.showwarning("Duplicate", f"{sid} has already checked in.")
            return
        if seat_input in used_seats:
            messagebox.showwarning("Seat taken", f"Seat {seat_input} is already used.")
            return

        _save_row(sid, full_name, seat_input)
        return

# ===================== GUI พื้นฐาน =====================
root = tk.Tk()
root.title("QR Attendance (Google Sheets)")

# ---- Base size for responsive ----
BASE_W, BASE_H = 1280, 720
root.geometry(f"{BASE_W}x{BASE_H}")
root.config(bg="#ffe5ec")

root.grid_rowconfigure(0, weight=1)
root.grid_columnconfigure(0, weight=1)

# ===== Main frame =====
frame = tk.Frame(root, bg=root["bg"])
frame.grid(row=0, column=0, sticky="nsew")

# ===== Fonts (pixel units; negative size) =====
BASE = {
    "title_px": 40,
    "info_px": 30,
    "entry_px": 30,
    "hint_px": 16,     # helper text size
    "result_px": 40,
    "wrap": 700,
    "pad_big": 14,
    "pad": 10,
    "ipadx": 40,
    "ipady": 10,
    "img_w": 504,
    "img_h": 218,
    "btn_padx": 16,
    "btn_pady": 10,
    "btn_width": 14,
}

title_font  = tkfont.Font(family="TH Sarabun New", size=-BASE["title_px"], weight="bold")
info_font   = tkfont.Font(family="TH Sarabun New", size=-BASE["info_px"], weight="bold")
entry_font  = tkfont.Font(family="TH Sarabun New", size=-BASE["entry_px"])
hint_font   = tkfont.Font(family="TH Sarabun New", size=-BASE["hint_px"])
result_font = tkfont.Font(family="TH Sarabun New", size=-BASE["result_px"], weight="bold")

# ===== Widgets =====
label = tk.Label(
    frame,
    text="General Chemistry Laboratory 2302113 2302163 2302178 — Semester 1, Academic Year 2568",
    font=title_font,
    bg=root["bg"],
    wraplength=BASE["wrap"],
)
label.pack(pady=BASE["pad_big"], expand=True)

# Image (optional)
try:
    original_img = Image.open("Chem chula.png")
except Exception:
    from PIL import Image as _Img
    original_img = _Img.new("RGBA", (BASE["img_w"], BASE["img_h"]), (255, 255, 255, 0))

img = ImageTk.PhotoImage(original_img.resize((BASE["img_w"], BASE["img_h"]), Image.LANCZOS))
image_label = tk.Label(frame, image=img, bg=frame["bg"])
image_label.pack(pady=BASE["pad"], expand=True)
image_label.image = img

info_label = tk.Label(
    frame,
    text="Scan CU Next barcode (camera) or enter Student ID / Seat and press Enter.",
    font=info_font,
    bg=root["bg"]
)
info_label.pack(pady=BASE["pad"], expand=True)

# === Inputs ===
entry_row = tk.Frame(frame, bg=frame["bg"])
entry_row.pack(pady=BASE["pad"], expand=True)

# Student ID
id_col = tk.Frame(entry_row, bg=frame["bg"])
id_col.pack(side="left", padx=12)
tk.Label(id_col, text="Student ID", font=entry_font, bg=frame["bg"]).pack()
entry = tk.Entry(id_col, font=entry_font, width=20, justify='center')
entry.pack(ipadx=BASE["ipadx"]//2, ipady=BASE["ipady"])
entry.focus()

# Seat
seat_col = tk.Frame(entry_row, bg=frame["bg"])
seat_col.pack(side="left", padx=12)
tk.Label(seat_col, text="Seat", font=entry_font, bg=frame["bg"]).pack()
seat_entry = tk.Entry(seat_col, font=entry_font, width=12, justify='center')
seat_entry.pack(ipadx=BASE["ipadx"]//3, ipady=BASE["ipady"])

# Buttons row
btn_row = tk.Frame(frame, bg=frame["bg"])
btn_row.pack(pady=BASE["pad"], expand=True)

check_btn = tk.Button(
    btn_row, text="Check-in", font=entry_font, command=lambda: check_in(),
    padx=BASE["btn_padx"], pady=BASE["btn_pady"], width=BASE["btn_width"]
)
check_btn.pack(side="left", padx=10)

# ===== Camera Button =====
def open_camera_window():
    CameraWindow(root, on_detect=lambda text: handle_camera_text(text))

def handle_camera_text(text: str):
    # ถูกเรียกเมื่อกล้องถอดรหัสบาร์โค้ดได้
    sid = extract_student_id(text)
    if sid:
        entry.delete(0, tk.END)
        entry.insert(0, sid)
        # auto-enter
        check_in()

cam_btn = tk.Button(
    btn_row, text="Open Camera", font=entry_font, command=open_camera_window,
    padx=BASE["btn_padx"], pady=BASE["btn_pady"], width=BASE["btn_width"]
)
cam_btn.pack(side="left", padx=10)

# Result
result_label = tk.Label(frame, text="", font=result_font, fg="navy", bg=frame["bg"])
result_label.pack(pady=BASE["pad"], expand=True)

# Bind Enter
entry.bind("<Return>", check_in)
seat_entry.bind("<Return>", check_in)

# ===================== Responsive (heartbeat) =====================
last_w, last_h = 0, 0

def apply_scale(scale: float):
    # fonts
    title_font.configure(size=-(max(10, int(BASE["title_px"]  * scale))))
    info_font.configure(size=-(max(10, int(BASE["info_px"]   * scale))))
    entry_font.configure(size=-(max(10, int(BASE["entry_px"] * scale))))
    hint_font.configure(size=-(max(8,  int(BASE["hint_px"]  * scale))))
    result_font.configure(size=-(max(10, int(BASE["result_px"]* scale))))

    # wrap/padding
    label.configure(wraplength=int(BASE["wrap"] * scale))
    pbig = int(BASE["pad_big"] * scale); p = int(BASE["pad"] * scale)

    label.pack_configure(pady=pbig)
    info_label.pack_configure(pady=p)
    image_label.pack_configure(pady=p)
    entry_row.pack_configure(pady=p)
    btn_row.pack_configure(pady=p)
    result_label.pack_configure(pady=p)

    # image
    target_w = max(1, int(round(BASE["img_w"] * scale)))
    target_h = max(1, int(round(BASE["img_h"] * scale)))
    resized = original_img.resize((target_w, target_h), Image.LANCZOS)
    ph = ImageTk.PhotoImage(resized)
    image_label.configure(image=ph)
    image_label.image = ph

def heartbeat():
    global last_w, last_h
    w = max(1, root.winfo_width())
    h = max(1, root.winfo_height())
    if w != last_w or h != last_h:
        scale = max(0.3, min(w/BASE_W, h/BASE_H))
        apply_scale(scale)
        last_w, last_h = w, h
    root.after(200, heartbeat)

root.after(50, heartbeat)

# ===================== กล้องด้วย OpenCV (Toplevel) =====================
class CameraWindow:
    def __init__(self, master, on_detect):
        self.on_detect = on_detect
        self.top = tk.Toplevel(master)
        self.top.title("Camera Scanner")
        self.top.protocol("WM_DELETE_WINDOW", self.on_close)

        self.label = tk.Label(self.top)
        self.label.pack()

        self.btn_frame = tk.Frame(self.top)
        self.btn_frame.pack(pady=5)
        tk.Button(self.btn_frame, text="Close", command=self.on_close).pack()

        # เปิดกล้อง
        self.cap = cv2.VideoCapture(0, cv2.CAP_DSHOW)  # อาจเปลี่ยน index ถ้ามีกล้องหลายตัว
        if not self.cap.isOpened():
            messagebox.showerror("Camera", "Cannot open camera.")
            self.top.destroy()
            return

        # ใช้ BarcodeDetector ของ OpenCV (contrib)
        try:
            self.bd = cv2.barcode_BarcodeDetector()
        except Exception:
            self.bd = None
        self.qr = cv2.QRCodeDetector()
        self.running = True
        self.last_text = ""

        self.update_frame()

    def decode_text(self, frame_bgr) -> str:
        # 1) OpenCV barcode module
        if self.bd is not None:
            ok, infos, types, corners = self.bd.detectAndDecode(frame_bgr)
            if ok and infos:
                t = (infos[0] or "").strip()
                if t:
                    return t
        # 2) fallback QR-only
        data, _, _ = self.qr.detectAndDecode(frame_bgr)
        return (data or "").strip()

    def update_frame(self):
        if not self.running:
            return
        ret, frame = self.cap.read()
        if ret:
            # ย่อภาพเพื่อแสดงผล
            disp = cv2.resize(frame, (800, 450))
            # ลองถอดรหัส
            txt = self.decode_text(frame)
            if txt and txt != self.last_text:
                self.last_text = txt
                # วนกลับไปกรอกช่องในหน้าหลัก
                try:
                    self.on_detect(txt)
                except Exception:
                    pass
            # แปลงเป็นรูป Tk
            rgb = cv2.cvtColor(disp, cv2.COLOR_BGR2RGB)
            imgtk = ImageTk.PhotoImage(image=Image.fromarray(rgb))
            self.label.configure(image=imgtk)
            self.label.image = imgtk
        self.top.after(30, self.update_frame)

    def on_close(self):
        self.running = False
        try:
            if self.cap and self.cap.isOpened():
                self.cap.release()
        except Exception:
            pass
        self.top.destroy()

# ===================== main loop =====================
root.mainloop()
